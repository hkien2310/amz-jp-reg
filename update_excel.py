import openpyxl

wb = openpyxl.load_workbook("AmazonJP.xlsx")

# 1. Tạo/Cập nhật sheet Gmails
if "Gmails" not in wb.sheetnames:
    wb.create_sheet("Gmails")
gmails_sheet = wb["Gmails"]

# Điền Header (dòng 1)
gmails_sheet.cell(row=1, column=1, value="email")
gmails_sheet.cell(row=1, column=2, value="password")
gmails_sheet.cell(row=1, column=3, value="otp_email")
gmails_sheet.cell(row=1, column=4, value="otp_pass")

# Điền dữ liệu của người dùng (dòng 2)
gmails_sheet.cell(row=2, column=1, value="zb.lackholez23@gmail.com")
gmails_sheet.cell(row=2, column=2, value="Zblackhole@2026") # Mật khẩu random để test
gmails_sheet.cell(row=2, column=3, value="zblackholez23@gmail.com")
# Strip out spaces from app password just in case imaplib hates spaces
app_pass = "tjzw ierz nuvf dhts".replace(" ", "").replace("\u2028", "").strip()
gmails_sheet.cell(row=2, column=4, value=app_pass)

# 2. Xóa sạch sheet Accounts cũ, chỉ để lại 1 dòng test
if "Accounts" in wb.sheetnames:
    acc_sheet = wb["Accounts"]
    # Clear dòng 2
    for r_idx in range(2, acc_sheet.max_row + 1):
        for c_idx in range(1, 10):
            acc_sheet.cell(row=r_idx, column=c_idx).value = None
            
    # Ghi dòng 2
    acc_sheet.cell(row=2, column=1, value=1)
    acc_sheet.cell(row=2, column=2, value="zb.lackholez23@gmail.com")
    acc_sheet.cell(row=2, column=5, value="pending")

wb.save("AmazonJP.xlsx")
print("Cập nhật AmazonJP.xlsx thành công!")
