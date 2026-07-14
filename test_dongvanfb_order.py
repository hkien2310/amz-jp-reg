import openpyxl
import requests
import json

wb = openpyxl.load_workbook("AmazonJP_test.xlsx")
sheet = wb["Accounts"]

for i in range(2, 5):
    email = sheet.cell(row=i, column=1).value
    if email == "jocastalanisophia5209@hotmail.com":
        password = sheet.cell(row=i, column=2).value
        # Assuming refresh_token is in some column. Actually wait, it's IMAP config, maybe the bot parses it from password?
        print("Found email. Wait, how is refresh token stored?")
        print(sheet.cell(row=i, column=8).value)
        
