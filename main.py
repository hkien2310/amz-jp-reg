import os
import sys
import time
import queue
import threading
import asyncio
import json
from datetime import datetime
import openpyxl
import customtkinter as ctk
from tkinter import filedialog, messagebox

# Import các module của chúng ta
from browser_helper import get_browser_path
from register_worker import register_amazon_account

# Cấu hình giao diện CustomTkinter
ctk.set_appearance_mode("System")  # Đồng bộ theo sáng/tối của hệ điều hành
ctk.set_default_color_theme("blue")  # Tông màu xanh chủ đạo

class ProxyManager:
    def __init__(self, max_usage, log_func):
        self.lock = threading.Lock()
        self.max_usage = max_usage
        self.log_func = log_func
        # proxies_usage = { "proxy_str": current_usage_count_from_excel }
        self.proxies_usage = {}
        # in_use = set of proxy_str currently being used by active workers
        self.in_use = set()
        
    def load_from_excel(self, ws):
        """Đọc danh sách proxy và số lần sử dụng từ Excel."""
        self.proxies_usage = {}
        for r_idx in range(2, ws.max_row + 1):
            proxy_val = ws.cell(row=r_idx, column=1).value
            if proxy_val:
                proxy_str = str(proxy_val).strip()
                usage_val = ws.cell(row=r_idx, column=3).value
                try:
                    usage_count = int(usage_val) if usage_val is not None else 0
                except ValueError:
                    usage_count = 0
                self.proxies_usage[proxy_str] = usage_count
                
    def get_proxy(self):
        """Lấy 1 proxy thoả mãn điều kiện: Usage < max_usage và chưa bị ai dùng lúc này."""
        with self.lock:
            for proxy, count in self.proxies_usage.items():
                if proxy not in self.in_use and count < self.max_usage:
                    self.in_use.add(proxy)
                    return proxy
            return None
            
    def release_proxy(self, proxy_str):
        """Giải phóng proxy sau khi worker dùng xong."""
        if not proxy_str:
            return
        with self.lock:
            if proxy_str in self.in_use:
                self.in_use.remove(proxy_str)
                
    def get_total_available(self):
        with self.lock:
            return sum(1 for count in self.proxies_usage.values() if count < self.max_usage)

class AmazonRegisterApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # Đường dẫn file cấu hình mặc định
        self.CONFIG_FILE = "config.json"
        
        # Cấu hình cửa sổ chính
        self.title("Amazon JP Auto Register Bot")
        self.geometry("900x700")
        self.minsize(850, 600)
        
        # Biến điều khiển dữ liệu và luồng
        self.excel_path = ctk.StringVar(value="")
        self.excel_lock = threading.Lock()
        self.pending_queue = queue.Queue()
        self.gui_otp_queue = queue.Queue()
        
        self.workers = []
        self.is_running = False
        self.total_accounts_to_run = 0
        self.accounts_processed = 0
        self.accounts_success = 0
        self.accounts_failed = 0
        
        # Tạo Giao diện UI
        self.create_widgets()
        
        # Nạp cấu hình tự động
        self.apply_saved_config()
        
        # Khởi tạo Proxy Manager
        config = self.load_config()
        self.proxy_manager = ProxyManager(
            max_usage=config.get("max_accounts_per_proxy", 2),
            log_func=self.write_log
        )
        
        # Bắt đầu vòng lặp kiểm tra yêu cầu OTP từ luồng phụ
        self.check_otp_requests()
        
    def create_widgets(self):
        # Thiết lập bố cục lưới (Grid Layout)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)  # Log box co giãn theo chiều dọc
        
        # --- FRAME 1: CHỌN FILE EXCEL ---
        self.file_frame = ctk.CTkFrame(self)
        self.file_frame.grid(row=0, column=0, padx=15, pady=10, sticky="ew")
        self.file_frame.grid_columnconfigure(1, weight=1)
        
        self.lbl_excel = ctk.CTkLabel(self.file_frame, text="File Excel:", font=ctk.CTkFont(weight="bold"))
        self.lbl_excel.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        self.entry_excel = ctk.CTkEntry(self.file_frame, textvariable=self.excel_path, state="readonly", placeholder_text="Chưa chọn file Excel dữ liệu...")
        self.entry_excel.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        
        self.btn_browse = ctk.CTkButton(self.file_frame, text="Chọn File", width=100, command=self.browse_excel)
        self.btn_browse.grid(row=0, column=2, padx=10, pady=10)
        
        # --- FRAME 2: CẤU HÌNH ---
        self.config_frame = ctk.CTkFrame(self)
        self.config_frame.grid(row=1, column=0, padx=15, pady=5, sticky="ew")
        self.config_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        # Đường dẫn Trình duyệt
        self.browser_path = ctk.StringVar(value="")
        self.lbl_browser = ctk.CTkLabel(self.config_frame, text="Đường dẫn Trình duyệt:", font=ctk.CTkFont(weight="bold"))
        self.lbl_browser.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        
        self.entry_browser = ctk.CTkEntry(self.config_frame, textvariable=self.browser_path, placeholder_text="Bỏ trống để tự tìm Chrome/Edge. Nhập path Cloak Browser nếu có")
        self.entry_browser.grid(row=0, column=1, columnspan=2, padx=10, pady=5, sticky="ew")
        
        self.btn_browse_browser = ctk.CTkButton(self.config_frame, text="Chọn", width=60, command=self.browse_browser)
        self.btn_browse_browser.grid(row=0, column=3, padx=10, pady=5, sticky="w")

        # Sheet chạy
        self.target_sheet = ctk.StringVar(value="Outlooks")
        self.lbl_target_sheet = ctk.CTkLabel(self.config_frame, text="Nguồn tài khoản:", font=ctk.CTkFont(weight="bold"))
        self.lbl_target_sheet.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.opt_target_sheet = ctk.CTkOptionMenu(self.config_frame, variable=self.target_sheet, values=["Outlooks", "Gmails", "Iclouds"])
        self.opt_target_sheet.grid(row=1, column=1, padx=10, pady=5, sticky="ew")

        # Số Worker
        self.lbl_workers = ctk.CTkLabel(self.config_frame, text="Số luồng (Workers): 1")
        self.lbl_workers.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        self.slider_workers = ctk.CTkSlider(self.config_frame, from_=1, to=5, number_of_steps=4, command=self.update_worker_label)
        self.slider_workers.set(1)
        self.slider_workers.grid(row=2, column=1, padx=10, pady=5, sticky="ew")
        
        # Giới hạn Account
        self.lbl_limit = ctk.CTkLabel(self.config_frame, text="Số Acc cần chạy:")
        self.lbl_limit.grid(row=1, column=2, padx=10, pady=5, sticky="e")
        self.entry_limit = ctk.CTkEntry(self.config_frame, width=80, placeholder_text="Tất cả")
        self.entry_limit.grid(row=1, column=3, padx=10, pady=5, sticky="w")
        
        # Checkbox Chạy ngầm, Proxy & Debug
        self.chk_headless = ctk.CTkCheckBox(self.config_frame, text="Chạy ngầm (Headless)")
        self.chk_headless.grid(row=3, column=0, padx=10, pady=10, sticky="w")
        
        self.chk_proxy = ctk.CTkCheckBox(self.config_frame, text="Sử dụng Proxy")
        self.chk_proxy.grid(row=3, column=1, padx=10, pady=10, sticky="w")
        
        self.chk_debug = ctk.CTkCheckBox(self.config_frame, text="Chế độ Debug (Lưu ảnh lỗi)")
        self.chk_debug.grid(row=3, column=2, columnspan=2, padx=10, pady=10, sticky="w")
        
        # --- FRAME 3: ĐIỀU KHIỂN & TIẾN TRÌNH ---
        self.control_frame = ctk.CTkFrame(self)
        self.control_frame.grid(row=2, column=0, padx=15, pady=5, sticky="ew")
        self.control_frame.grid_columnconfigure(2, weight=1)
        
        self.btn_start = ctk.CTkButton(self.control_frame, text="BẮT ĐẦU CHẠY", fg_color="#10b981", hover_color="#059669", font=ctk.CTkFont(weight="bold"), command=self.btn_start_click)
        self.btn_start.grid(row=0, column=0, padx=15, pady=15, ipadx=10, ipady=5)
        
        self.btn_stop = ctk.CTkButton(self.control_frame, text="DỪNG LẠI", fg_color="#ef4444", hover_color="#dc2626", font=ctk.CTkFont(weight="bold"), state="disabled", command=self.btn_stop_click)
        self.btn_stop.grid(row=0, column=1, padx=15, pady=15, ipadx=10, ipady=5)
        
        # Tiến trình và thống kê
        self.stats_frame = ctk.CTkFrame(self.control_frame, fg_color="transparent")
        self.stats_frame.grid(row=0, column=2, padx=10, pady=10, sticky="nsew")
        self.stats_frame.grid_columnconfigure(0, weight=1)
        
        self.lbl_stats = ctk.CTkLabel(self.stats_frame, text="Sẵn sàng chạy | Thành công: 0 | Thất bại: 0", font=ctk.CTkFont(size=12))
        self.lbl_stats.grid(row=0, column=0, sticky="w")
        
        self.progress_bar = ctk.CTkProgressBar(self.stats_frame, width=300)
        self.progress_bar.set(0)
        self.progress_bar.grid(row=1, column=0, pady=5, sticky="ew")
        
        # --- FRAME 4: LOG CONSOLE & WORKERS STATUS ---
        self.display_frame = ctk.CTkFrame(self)
        self.display_frame.grid(row=3, column=0, padx=15, pady=10, sticky="nsew")
        self.display_frame.grid_columnconfigure(0, weight=3)  # Log chiếm 3 phần
        self.display_frame.grid_columnconfigure(1, weight=1)  # Status chiếm 1 phần
        self.display_frame.grid_rowconfigure(0, weight=1)
        
        # Log Box
        self.log_textbox = ctk.CTkTextbox(self.display_frame, font=ctk.CTkFont(family="Courier", size=12))
        self.log_textbox.grid(row=0, column=0, padx=(10, 5), pady=(10, 5), sticky="nsew")
        self.log_textbox.insert("end", "=== HỆ THỐNG ĐÃ SẴN SÀNG ===\nVui lòng chọn file Excel và cấu hình để bắt đầu.\n")
        self.log_textbox.configure(state="disabled")
        
        # Nút copy log
        self.btn_copy_log = ctk.CTkButton(self.display_frame, text="Sao chép toàn bộ Log", font=ctk.CTkFont(size=11), width=150, command=self.copy_all_logs)
        self.btn_copy_log.grid(row=1, column=0, padx=(10, 5), pady=(0, 10), sticky="w")
        
        # Status Box (Trạng thái Workers)
        self.status_frame = ctk.CTkScrollableFrame(self.display_frame, label_text="Trạng thái các luồng")
        self.status_frame.grid(row=0, column=1, rowspan=2, padx=(5, 10), pady=10, sticky="nsew")
        
        self.worker_labels = {}
        
    def update_worker_label(self, val):
        self.lbl_workers.configure(text=f"Số luồng (Workers): {int(val)}")
        
    def write_log(self, text):
        """Ghi log an toàn từ bất kỳ luồng nào vào Log Box."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_textbox.configure(state="normal")
        self.log_textbox.insert("end", f"[{timestamp}] {text}\n")
        self.log_textbox.see("end")
        self.log_textbox.configure(state="disabled")
        
    def browse_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.excel_path.set(file_path)
            self.write_log(f"Đã chọn file Excel: {os.path.basename(file_path)}")
            # Kiểm tra sơ bộ các sheet
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                missing = [s for s in ["Accounts", "Emails", "Proxies"] if s not in sheets]
                if missing:
                    self.write_log(f"CẢNH BÁO: File Excel thiếu các sheet bắt buộc: {', '.join(missing)}")
                    messagebox.showwarning("Thiếu Sheet", f"File Excel phải chứa đủ các sheet: Accounts, Emails, Proxies.\n\nThiếu: {', '.join(missing)}")
                else:
                    self.write_log("Kiểm tra file Excel: HỢP LỆ (Đầy đủ sheet)")
            except Exception as e:
                self.write_log(f"LỖI đọc file Excel: {str(e)}")
                messagebox.showerror("Lỗi file", f"Không thể đọc file Excel:\n{str(e)}")

    def browse_browser(self):
        if sys.platform == "darwin":
            file_path = filedialog.askopenfilename(title="Chọn trình duyệt (Chrome/Brave/Chromium)")
        else:
            file_path = filedialog.askopenfilename(title="Chọn trình duyệt", filetypes=[("Executables", "*.exe"), ("All Files", "*.*")])
        if file_path:
            self.browser_path.set(file_path)
            self.write_log(f"Đã cập nhật đường dẫn trình duyệt tuỳ chỉnh.")

                
    def check_otp_requests(self):
        """Hàm kiểm tra định kỳ xem có luồng nào yêu cầu popup OTP không."""
        try:
            while not self.gui_otp_queue.empty():
                request = self.gui_otp_queue.get_nowait()
                email_addr = request["email"]
                
                # Mở hộp thoại nhập OTP trên luồng chính
                dialog = ctk.CTkInputDialog(text=f"Mã xác minh Amazon OTP gửi tới:\n{email_addr}\n\nVui lòng nhập mã OTP 6 chữ số:", title="Nhập OTP Thủ Công")
                code = dialog.get_input()
                
                # Gán kết quả và kích hoạt luồng phụ tiếp tục chạy
                request["code"] = code
                request["event"].set()
                self.gui_otp_queue.task_done()
        except Exception as e:
            pass
            
        # Kiểm tra lại sau 500ms
        self.after(500, self.check_otp_requests)
        
    def start_bot(self):
        # 1. Kiểm tra file Excel
        path = self.excel_path.get()
        if not path or not os.path.exists(path):
            messagebox.showerror("Thiếu file", "Vui lòng chọn file Excel dữ liệu trước khi chạy!")
            return
            
        # 2. Đọc cấu hình từ UI
        try:
            limit_val = self.entry_limit.get().strip()
            limit = int(limit_val) if limit_val else None
        except ValueError:
            messagebox.showerror("Cấu hình lỗi", "Số lượng Acc cần chạy phải là số nguyên hợp lệ hoặc để trống để chạy toàn bộ!")
            return
            
        num_workers = int(self.slider_workers.get())
        use_proxy = self.chk_proxy.get() == 1
        headless = self.chk_headless.get() == 1
        debug_mode = self.chk_debug.get() == 1
        
        # Lưu cấu hình lại
        self.save_config()
        
        # 3. Nạp dữ liệu từ Excel vào Queue
        self.write_log("Đang phân tích dữ liệu Excel...")
        success_init = self.load_pending_accounts_to_queue(path, use_proxy, limit)
        
        if not success_init:
            return
            
        # 4. Cấu hình trạng thái chạy
        self.is_running = True
        self.accounts_processed = 0
        self.accounts_success = 0
        self.accounts_failed = 0
        self.progress_bar.set(0)
        self.update_stats_label()
        
        # Vô hiệu hóa nút và cấu hình khi chạy
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.btn_browse.configure(state="disabled")
        self.slider_workers.configure(state="disabled")
        self.entry_limit.configure(state="disabled")
        self.chk_headless.configure(state="disabled")
        self.chk_proxy.configure(state="disabled")
        self.chk_debug.configure(state="disabled")
        
        # Dọn dẹp status box và tạo nhãn hiển thị cho các worker
        for widget in self.status_frame.winfo_children():
            widget.destroy()
            
        self.worker_labels = {}
        for w_id in range(1, num_workers + 1):
            lbl = ctk.CTkLabel(self.status_frame, text=f"Worker {w_id}: Chờ chạy...", anchor="w")
            lbl.pack(fill="x", padx=5, pady=2)
            self.worker_labels[w_id] = lbl
            
        # Lưu cache excel path để dùng trong các luồng worker (do get() của StringVar không thread-safe)
        self.cached_excel_path = self.excel_path.get()
            
        # 5. Khởi chạy luồng chính điều phối
        browser_path_val = self.browser_path.get().strip()
        self.coordinator_thread = threading.Thread(target=self.run_coordinator, args=(num_workers, headless, debug_mode, browser_path_val))
        self.coordinator_thread.daemon = True
        self.coordinator_thread.start()
        
    def btn_start_click(self):
        if self.is_running:
            return
            
        import config
        config.GLOBAL_STOP = False
        
        excel_path_val = self.excel_path.get().strip()
        if not excel_path_val or not os.path.exists(excel_path_val):
            messagebox.showerror("Lỗi", "Vui lòng chọn file Excel hợp lệ!")
            return
            
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.btn_browse.configure(state="disabled")
        self.slider_workers.configure(state="disabled")
        self.entry_limit.configure(state="disabled")
        self.chk_headless.configure(state="disabled")
        self.chk_proxy.configure(state="disabled")
        self.chk_debug.configure(state="disabled")
        
        self.accounts_processed = 0
        self.accounts_success = 0
        self.accounts_failed = 0
        
        self.write_log("Đang phân tích dữ liệu Excel...")
        
        try:
            limit_val = int(self.entry_limit.get().strip())
        except ValueError:
            limit_val = None
            
        use_proxy = self.chk_proxy.get() == 1
        headless = self.chk_headless.get() == 1
        debug_mode = self.chk_debug.get() == 1
        
        success = self.load_pending_accounts_to_queue(excel_path_val, use_proxy, limit_val)
        if not success:
            self.reset_gui_after_run()
            return
            
        self.is_running = True
        self.update_stats_label()
        
        num_workers = int(self.slider_workers.get())
        
        # Tạo lại status worker labels
        for widget in self.status_frame.winfo_children():
            widget.destroy()
            
        self.worker_labels = {}
        for w_id in range(1, num_workers + 1):
            lbl = ctk.CTkLabel(self.status_frame, text=f"Worker {w_id}: Chờ chạy...", anchor="w")
            lbl.pack(fill="x", padx=5, pady=2)
            self.worker_labels[w_id] = lbl
            
        self.cached_excel_path = self.excel_path.get()
            
        browser_path_val = self.browser_path.get().strip()
        self.coordinator_thread = threading.Thread(target=self.run_coordinator, args=(num_workers, headless, debug_mode, browser_path_val))
        self.coordinator_thread.daemon = True
        self.coordinator_thread.start()
        
    def btn_stop_click(self):
        if not self.is_running:
            return
            
        import config
        config.GLOBAL_STOP = True
        self.is_running = False
        self.write_log("Đang yêu cầu dừng các luồng... Vui lòng chờ!")
        self.btn_stop.configure(state="disabled")
        
    def load_pending_accounts_to_queue(self, excel_path, use_proxy, limit):
        """Đọc file Excel và đưa các account 'pending' vào hàng đợi."""
        try:
            wb = openpyxl.load_workbook(excel_path)
            target_sheet_name = self.target_sheet.get()
            if target_sheet_name not in wb.sheetnames:
                self.write_log(f"CẢNH BÁO: Không tìm thấy sheet nguồn '{target_sheet_name}' trong file Excel!")
                messagebox.showwarning("Lỗi Sheet", f"Không tìm thấy sheet '{target_sheet_name}'!")
                return False
                
            # Đọc Proxies
            if use_proxy:
                if 'Proxies' in wb.sheetnames:
                    proxies_sheet = wb['Proxies']
                    self.proxy_manager.load_from_excel(proxies_sheet)
                if len(self.proxy_manager.proxies_usage) == 0:
                    self.write_log("CẢNH BÁO: Bật chạy proxy nhưng sheet Proxies không có dữ liệu!")
                    
            # Đọc danh sách cần chạy từ sheet nguồn (Outlooks, Gmails, Iclouds)
            source_sheet = wb[target_sheet_name]
            pending_tasks = []
            
            for r_idx in range(2, source_sheet.max_row + 1):
                email_addr = source_sheet.cell(row=r_idx, column=1).value
                email_pass = source_sheet.cell(row=r_idx, column=2).value
                otp_email = source_sheet.cell(row=r_idx, column=3).value
                otp_pass = source_sheet.cell(row=r_idx, column=4).value
                
                # Xác định cột Status tuỳ theo sheet
                if target_sheet_name == "Outlooks":
                    status = source_sheet.cell(row=r_idx, column=2).value # Cột 2 lưu status của Outlooks
                else:
                    status = source_sheet.cell(row=r_idx, column=5).value # Cột 5 cho Gmails/Iclouds
                
                refresh_token = ""
                client_id = ""
                
                # Check nếu Outlooks vẫn dùng định dạng gộp cũ email|pass|token|client_id thì tách ra
                if email_addr and "|" in str(email_addr):
                    parts = str(email_addr).split("|")
                    if len(parts) >= 2:
                        email_addr = parts[0].strip()
                        # Đối với định dạng cũ, pass có thể nằm ở cột 2 hoặc trong chuỗi gộp
                        email_pass = parts[1].strip()
                        refresh_token = parts[2].strip() if len(parts) >= 3 else ""
                        client_id = parts[3].strip() if len(parts) >= 4 else ""
                        otp_email = None
                        otp_pass = None

                if not email_addr:
                    continue
                    
                # Chạy lại cả Email bị Fail và Pending
                is_pending = not status or str(status).strip().lower() in ["pending", "none", "", "fail", "failed"]
                if is_pending:
                    assigned_proxy = None
                    email_row_idx = r_idx
                    pending_tasks.append((r_idx, str(email_addr).strip(), str(email_pass).strip() if email_pass else "", email_row_idx, assigned_proxy, refresh_token, client_id, str(otp_email).strip() if otp_email else None, str(otp_pass).strip() if otp_pass else None, target_sheet_name))
                        
            wb.close()
            
            # Áp dụng giới hạn số lượng chạy
            if limit and len(pending_tasks) > limit:
                pending_tasks = pending_tasks[:limit]
                
            self.total_accounts_to_run = len(pending_tasks)
            self.write_log(f"Tìm thấy {self.total_accounts_to_run} accounts ở trạng thái 'pending' cần xử lý.")
            
            if self.total_accounts_to_run == 0:
                self.write_log("Không có account nào cần xử lý.")
                messagebox.showinfo("Hoàn thành", "Không tìm thấy account nào ở trạng thái 'pending'!")
                return False
                
            # Đưa vào Queue của ứng dụng
            while not self.pending_queue.empty():
                self.pending_queue.get()
                
            for task in pending_tasks:
                self.pending_queue.put(task)
                
            return True
            
        except Exception as e:
            self.write_log(f"LỖI load Excel vào Queue: {str(e)}")
            return False
            
    def update_stats_label(self):
        percentage = 0
        if self.total_accounts_to_run > 0:
            percentage = self.accounts_processed / self.total_accounts_to_run
            
        self.progress_bar.set(percentage)
        self.lbl_stats.configure(
            text=f"Tiến trình: {self.accounts_processed} / {self.total_accounts_to_run} | Thành công: {self.accounts_success} | Thất bại: {self.accounts_failed}"
        )
        
    def run_coordinator(self, num_workers, headless, debug_mode, browser_path_val):
        self.write_log(f"Bắt đầu khởi chạy {num_workers} workers đăng ký song song...")
        
        self.workers = []
        for w_id in range(1, num_workers + 1):
            t = threading.Thread(target=self.worker_loop, args=(w_id, headless, debug_mode, browser_path_val))
            t.daemon = True
            t.start()
            self.workers.append(t)
            
        # Đợi các worker kết thúc
        for t in self.workers:
            t.join()
            
        self.is_running = False
        self.write_log("=== TẤT CẢ LUỒNG ĐÃ DỪNG HOẠT ĐỘNG ===")
        
        # Khôi phục trạng thái GUI trên luồng chính
        self.after(0, self.reset_gui_after_run)
        
    def reset_gui_after_run(self):
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self.btn_browse.configure(state="normal")
        self.slider_workers.configure(state="normal")
        self.entry_limit.configure(state="normal")
        self.chk_headless.configure(state="normal")
        self.chk_proxy.configure(state="normal")
        self.chk_debug.configure(state="normal")
        
        # Thông báo hoàn tất
        messagebox.showinfo("Kết thúc", f"Chương trình chạy xong!\n\nThành công: {self.accounts_success}\nThất bại: {self.accounts_failed}")
        
    def get_otp_from_gui_popup(self, email_addr):
        """Hàm callback gọi từ luồng worker để gửi yêu cầu popup OTP lên luồng chính."""
        request = {
            "email": email_addr,
            "event": threading.Event(),
            "code": None
        }
        self.gui_otp_queue.put(request)
        # Chờ luồng chính giải quyết popup và set event
        request["event"].wait()
        return request["code"]
        
    def update_worker_status_gui(self, w_id, status_text):
        """Cập nhật label trạng thái worker lên giao diện."""
        self.after(0, lambda: self.worker_labels[w_id].configure(text=f"Worker {w_id}: {status_text}"))
        
    def worker_loop(self, worker_id, headless, debug_mode, browser_path_val):
        def worker_log(w_id, msg):
            self.write_log(f"Worker {w_id}: {msg}")
            
        self.update_worker_status_gui(worker_id, "Sẵn sàng")
        
        # Biến loop kiểm tra liên tục hàng đợi
        while self.is_running:
            try:
                # Lấy công việc từ Queue (timeout 1s để dễ thoát khi click Stop)
                task = self.pending_queue.get(timeout=1)
            except queue.Empty:
                break
                
            accounts_row_idx, email_addr, email_pass, email_row_idx, _, refresh_token, client_id, otp_email, otp_pass, target_sheet_name = task
            
            # Cấp phát proxy nếu cần
            proxy_str = None
            if self.chk_proxy.get() == 1:
                proxy_str = self.proxy_manager.get_proxy()
                if not proxy_str:
                    worker_log(worker_id, "Không còn proxy khả dụng (đạt max_accounts hoặc đang bận). Trả lại account vào hàng đợi và dừng worker.")
                    # Trả lại task
                    self.pending_queue.put(task)
                    break
                    
            self.update_worker_status_gui(worker_id, f"Đăng ký {email_addr}")
            
            start_time = time.time()
            
            # Chạy tự động hóa đăng ký trên Amazon
            # Gọi asyncio run vì Playwright là thư viện async
            try:
                status, phone, registered_name, registered_pass, error_msg = asyncio.run(
                    register_amazon_account(
                        worker_id,
                        email_addr,
                        email_pass,
                        proxy_str,
                        headless,
                        debug_mode,
                        self.get_otp_from_gui_popup,
                        worker_log,
                        refresh_token=refresh_token,
                        client_id=client_id,
                        custom_browser_path=browser_path_val,
                        otp_email=otp_email,
                        otp_pass=otp_pass
                    )
                )
            except Exception as e:
                status, phone, registered_name, registered_pass, error_msg = "FAILED", None, None, None, str(e)
                
            elapsed_time = int(time.time() - start_time)
            
            # Cập nhật kết quả vào file Excel (Thread Safe qua Lock)
            with self.excel_lock:
                try:
                    wb = openpyxl.load_workbook(self.cached_excel_path)
                    
                    # 1. Thêm kết quả vào sheet Accounts (sheet kết quả chính)
                    acc_sheet = wb['Accounts']
                    
                    # Tìm xem email này đã tồn tại trong sheet Accounts chưa để ghi đè, nếu chưa thì append
                    existing_row = None
                    for r in range(2, acc_sheet.max_row + 1):
                        if str(acc_sheet.cell(row=r, column=2).value).strip() == str(email_addr).strip():
                            existing_row = r
                            break
                            
                    if existing_row:
                        target_row = existing_row
                    else:
                        target_row = acc_sheet.max_row + 1
                        acc_sheet.cell(row=target_row, column=1, value=target_row - 1) # STT
                        acc_sheet.cell(row=target_row, column=2, value=email_addr) # Email
                        
                    acc_sheet.cell(row=target_row, column=3, value=registered_name if status == "SUCCESS" else None)
                    acc_sheet.cell(row=target_row, column=4, value=registered_pass if status == "SUCCESS" else None)
                    acc_sheet.cell(row=target_row, column=5, value=status)
                    acc_sheet.cell(row=target_row, column=6, value=phone)
                    acc_sheet.cell(row=target_row, column=7, value=proxy_str if proxy_str else "direct")
                    acc_sheet.cell(row=target_row, column=8, value=elapsed_time)
                    acc_sheet.cell(row=target_row, column=9, value=error_msg)
                    acc_sheet.cell(row=target_row, column=10, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    
                    # 2. Cập nhật trạng thái vào sheet nguồn (Outlooks/Gmails/Iclouds)
                    if target_sheet_name in wb.sheetnames:
                        source_sheet = wb[target_sheet_name]
                        if target_sheet_name == "Outlooks":
                            # Ghi status vào cột 2 (giống Emails cũ)
                            source_sheet.cell(row=email_row_idx, column=2, value=status)
                        else:
                            # Ghi status vào cột 5 của Iclouds/Gmails
                            source_sheet.cell(row=email_row_idx, column=5, value=status)
                    if status == "SUCCESS" and proxy_str:
                        proxies_sheet = wb['Proxies']
                        for r_idx in range(2, proxies_sheet.max_row + 1):
                            if str(proxies_sheet.cell(row=r_idx, column=1).value).strip() == proxy_str:
                                current_val = proxies_sheet.cell(row=r_idx, column=3).value
                                try:
                                    count = int(current_val) if current_val is not None else 0
                                except:
                                    count = 0
                                proxies_sheet.cell(row=r_idx, column=3, value=count + 1)
                                # Cập nhật cache trong ProxyManager
                                self.proxy_manager.proxies_usage[proxy_str] = count + 1
                                break
                                
                    wb.save(self.cached_excel_path)
                    wb.close()
                    
                    worker_log(worker_id, f"Ghi kết quả Excel thành công ({status})")
                except Exception as ex:
                    worker_log(worker_id, f"LỖI lưu kết quả vào Excel: {str(ex)}")
            
            # Cập nhật thống kê tiến trình lên GUI
            self.accounts_processed += 1
            if status == "SUCCESS":
                self.accounts_success += 1
            else:
                self.accounts_failed += 1
                
            self.after(0, self.update_stats_label)
            
            # Giải phóng proxy
            if proxy_str:
                self.proxy_manager.release_proxy(proxy_str)
            
            self.pending_queue.task_done()
            time.sleep(2)  # Nghỉ ngắn giữa các lần xử lý để tránh spam
            
        self.update_worker_status_gui(worker_id, "Đã dừng")

    def copy_all_logs(self):
        """Sao chép toàn bộ nội dung trong Log Box vào clipboard."""
        try:
            log_text = self.log_textbox.get("1.0", "end-1c")
            self.clipboard_clear()
            self.clipboard_append(log_text)
            self.update()
            self.write_log("Đã sao chép toàn bộ log vào clipboard!")
            messagebox.showinfo("Thành công", "Đã sao chép toàn bộ log vào clipboard!")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể sao chép log: {str(e)}")

    def load_config(self):
        """Đọc cấu hình từ file config.json nếu có."""
        defaults = {
            "excel_path": "",
            "browser_path": "",
            "target_sheet": "Outlooks",
            "workers": 1,
            "limit": "",
            "headless": False,
            "use_proxy": True,
            "debug_mode": True,
            "max_accounts_per_proxy": 2
        }
        if os.path.exists(self.CONFIG_FILE):
            try:
                with open(self.CONFIG_FILE, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    defaults.update(config)
            except Exception as e:
                self.write_log(f"Lỗi khi đọc file cấu hình: {e}")
        return defaults

    def apply_saved_config(self):
        """Áp dụng các giá trị cấu hình vào giao diện GUI."""
        config = self.load_config()
        self.excel_path.set(config.get("excel_path", ""))
        self.browser_path.set(config.get("browser_path", ""))
        self.target_sheet.set(config.get("target_sheet", "Outlooks"))
        self.slider_workers.set(config.get("workers", 1))
        self.update_worker_label(config.get("workers", 1))
        
        self.entry_limit.delete(0, "end")
        self.entry_limit.insert(0, config.get("limit", ""))
        
        if config.get("headless", False):
            self.chk_headless.select()
        else:
            self.chk_headless.deselect()
            
        if config.get("use_proxy", True):
            self.chk_proxy.select()
        else:
            self.chk_proxy.deselect()
            
        if config.get("debug_mode", True):
            self.chk_debug.select()
        else:
            self.chk_debug.deselect()

    def save_config(self):
        """Lưu cấu hình hiện tại của GUI vào file config.json."""
        # Giữ lại các cấu hình ngầm định không nằm trên GUI
        config = self.load_config()
        config.update({
            "excel_path": self.excel_path.get(),
            "browser_path": self.browser_path.get(),
            "target_sheet": self.target_sheet.get(),
            "workers": int(self.slider_workers.get()),
            "limit": self.entry_limit.get().strip(),
            "headless": self.chk_headless.get() == 1,
            "use_proxy": self.chk_proxy.get() == 1,
            "debug_mode": self.chk_debug.get() == 1,
        })
        try:
            with open(self.CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            self.write_log("Đã tự động lưu cấu hình thông số.")
        except Exception as e:
            self.write_log(f"Lỗi khi ghi file cấu hình: {e}")

if __name__ == "__main__":
    # Fix lỗi chạy asyncio loop trong Thread phụ trên một số nền tảng
    import platform
    if platform.system() == "Windows":
        import asyncio
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        
    app = AmazonRegisterApp()
    app.mainloop()
